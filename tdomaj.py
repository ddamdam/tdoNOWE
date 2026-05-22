import streamlit as st
import pandas as pd
import io
import re
import os
from pathlib import Path
from typing import Dict, Optional, List, Tuple

import pymupdf as fitz
import openpyxl
from openpyxl.utils import get_column_letter

# Szablon Excel: wiersz 2 = źródło (architekt/marketing/ocr), wiersz 3 = nagłówki, dane od wiersza 4.
# Program uzupełnia WYŁĄCZNIE kolumny oznaczone jako "ocr".
EXCEL_COLUMNS: List[Dict] = [
    {"idx": 1, "source": "architekt", "header": "Numer lokalu z dokumentacji budowlanej"},
    {"idx": 2, "source": "architekt", "header": "Powierzchnia z dokumentacji budowlanej dla GW (m2) (z garażami)\n*UWAGA: powierzchnia do zweryfikowania po otrzymaniu proj. budowlanego"},
    {"idx": 3, "source": "architekt", "header": "Powierzchnia z dokumentacji budowlanej dla GW (m2) + pow. schodów"},
    {"idx": 4, "source": "marketing", "header": "OSIEDLE"},
    {"idx": 5, "source": "marketing", "header": "Miasto/Miejscowość"},
    {"idx": 6, "source": "marketing", "header": "Dzielnica"},
    {"idx": 7, "source": "marketing", "header": "Numer budynku"},
    {"idx": 8, "source": "ocr", "header": "NUMER LOKALU"},
    {"idx": 9, "source": "marketing", "header": "Nazwa robocza typu"},
    {"idx": 10, "source": "marketing", "header": "OPIS LOKALU "},
    {"idx": 11, "source": "architekt", "header": "ETAP"},
    {"idx": 12, "source": "marketing", "header": "URL "},
    {"idx": 13, "source": "architekt", "header": "TYP"},
    {"idx": 14, "source": "ocr", "header": "EKSPOZYCJA "},
    {"idx": 15, "source": "ocr", "header": "POWIERZCHNIA UZYTKOWA PARTER"},
    {"idx": 16, "source": "ocr", "header": "POWIERZCHNIA UZYTKOWA PIĘTRO "},
    {"idx": 17, "source": "ocr", "header": "POWIERZCHNIA UZYTKOWA PIETRO 2"},
    {"idx": 18, "source": "ocr", "header": "Powierzchnia pod ściankami i schodów Parter "},
    {"idx": 19, "source": "ocr", "header": "Powierzchnia pod ściankami  i schodów  Piętro  "},
    {"idx": 20, "source": "ocr", "header": "Powierzchnia pod ściankami  i schodów  Piętro  2"},
    {"idx": 21, "source": "ocr", "header": "łączna powierzchnia parter"},
    {"idx": 22, "source": "ocr", "header": "łączna powierzchnia piętro"},
    {"idx": 23, "source": "ocr", "header": "łączna powierzchnia piętro 2"},
    {"idx": 24, "source": "ocr", "header": "POWIERZCHNIA ŁĄCZNIE"},
    {"idx": 25, "source": "ocr", "header": "POWIERZCHNIA UŻYTKOWA ŁĄCZNIE"},
    {"idx": 26, "source": "ocr", "header": "LICZBA POKOI "},
    {"idx": 27, "source": "ocr", "header": "LICZBA KONDYGNACJI "},
    {"idx": 29, "source": "ocr", "header": "OGRÓD"},
    {"idx": 30, "source": "ocr", "header": "PODDASZE"},
    {"idx": 31, "source": "ocr", "header": "Wysokość PARTERU"},
    {"idx": 32, "source": "ocr", "header": "Wysokość PIĘTRA"},
    {"idx": 33, "source": "ocr", "header": "Wysokość PODDASZA"},
    {"idx": 34, "source": "ocr", "header": "TARAS "},
    {"idx": 35, "source": "ocr", "header": "LOGGIA"},
    {"idx": 36, "source": "ocr", "header": "BALKON"},
    {"idx": 37, "source": "marketing", "header": "GARAŻ 1"},
    {"idx": 38, "source": "marketing", "header": "CENA GARAŻU"},
    {"idx": 39, "source": "marketing", "header": "GARAŻ 2"},
    {"idx": 40, "source": "marketing", "header": "CENA GARAŻU"},
    {"idx": 41, "source": "marketing", "header": "MIEJSCE POSTOJOWE 1"},
    {"idx": 42, "source": "marketing", "header": "CENA MP"},
    {"idx": 43, "source": "marketing", "header": "MIEJSCE POSTOJOWE 2"},
    {"idx": 44, "source": "marketing", "header": "CENA MP"},
    {"idx": 45, "source": "marketing", "header": "NAROŻNY "},
    {"idx": 46, "source": "ocr", "header": "KUCHNIA"},
    {"idx": 47, "source": "ocr", "header": "ANEKS KUCHENNY "},
    {"idx": 48, "source": "ocr", "header": "GARDEROBA"},
    {"idx": 49, "source": "ocr", "header": "KOMINEK "},
    {"idx": 50, "source": "architekt", "header": "Okno w kuchni lub aneksie"},
    {"idx": 51, "source": "ocr", "header": "Schowek"},
    {"idx": 52, "source": "ocr", "header": "Spiżarnia"},
    {"idx": 53, "source": "ocr", "header": "Pomieszczenie gospodarcze"},
    {"idx": 54, "source": "architekt", "header": "Lokal połączony z garażem"},
    {"idx": 55, "source": "ocr", "header": "Druga łazienka z prysznicem"},
    {"idx": 56, "source": "ocr", "header": "Gabinet"},
    {"idx": 57, "source": "", "header": "Balkon"},
    {"idx": 58, "source": "", "header": "Loggia"},
    {"idx": 59, "source": "", "header": "Taras"},
    {"idx": 60, "source": "marketing", "header": "CENA "},
    {"idx": 61, "source": "marketing", "header": "CENA M2 "},
    {"idx": 62, "source": "marketing", "header": "FEATURES"},
]


TEMPLATE_BASENAME = "ostateczny - kopia.xlsx"


def _script_dir() -> Path:
    return Path(__file__).resolve().parent


def create_excel_template(path: Path) -> None:
    """Tworzy szablon (wiersz 2 = źródło, wiersz 3 = nagłówki, dane od wiersza 4)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arkusz1"
    for col_def in EXCEL_COLUMNS:
        ws.cell(2, col_def["idx"], col_def["source"])
        ws.cell(3, col_def["idx"], col_def["header"])
    wb.save(path)


def ensure_excel_template() -> str:
    """Szablon obok skryptu — tworzy plik, jeśli go nie ma (Streamlit Cloud)."""
    script_dir = _script_dir()
    for name in (TEMPLATE_BASENAME, "ostateczny — kopia.xlsx"):
        path = script_dir / name
        if path.is_file():
            return str(path)
    matches = sorted(script_dir.glob("ostateczny*kopia*.xlsx"))
    if matches:
        return str(matches[0])
    path = script_dir / TEMPLATE_BASENAME
    create_excel_template(path)
    return str(path)


def find_excel_template() -> str:
    return ensure_excel_template()


def build_full_row(result: Dict[str, str]) -> List[str]:
    """Pełny wiersz danych: OCR wypełnione, marketing/architekt puste."""
    row = []
    for col_def in EXCEL_COLUMNS:
        if col_def["source"] == "ocr":
            row.append(result.get(col_def["header"], ""))
        else:
            row.append("")
    return row


def results_to_dataframe(results: List[Dict[str, str]]) -> pd.DataFrame:
    columns = [
        f"{get_column_letter(c['idx'])} | {c['header'].replace(chr(10), ' ')}"
        for c in EXCEL_COLUMNS
    ]
    rows = [build_full_row(result) for result in results]
    return pd.DataFrame(rows, columns=columns)


def export_results_to_excel(results: List[Dict[str, str]]) -> bytes:
    wb = openpyxl.load_workbook(ensure_excel_template())
    ws = wb["Arkusz1"] if "Arkusz1" in wb.sheetnames else wb.active
    if ws.max_row >= 4:
        ws.delete_rows(4, ws.max_row - 3)

    for row_idx, result in enumerate(results, start=4):
        for col_def, value in zip(EXCEL_COLUMNS, build_full_row(result)):
            ws.cell(row_idx, col_def["idx"], value if value != "" else None)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


class ApartmentPlanOCR:
    def __init__(self):
        self.columns = [c["header"] for c in EXCEL_COLUMNS]

    def empty_row(self) -> Dict[str, str]:
        return {header: "" for header in self.columns}

    def _safe_float_sum(self, *args: str) -> str:
        total = 0.0
        for val in args:
            if isinstance(val, str) and val.strip():
                try:
                    total += float(val.replace(",", ".").strip())
                except (ValueError, TypeError):
                    continue
        return f"{total:.2f}" if total > 0 else ""

    def extract_text_from_pdf(self, pdf_file_obj: io.BytesIO) -> Tuple[str, int]:
        full_text = ""
        page_count = 0
        try:
            pdf_bytes = pdf_file_obj.getvalue()
            with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
                page_count = doc.page_count
                for i, page in enumerate(doc):
                    blocks = page.get_text("blocks", sort=True)
                    page_text = " ".join(
                        [block[4].strip().replace("\n", " ") for block in blocks if block[4].strip()]
                    )
                    full_text += f"\n--- STRONA {i + 1} ---\n" + page_text
            return full_text, page_count
        except Exception as e:
            st.error(f"Błąd podczas odczytu pliku PDF za pomocą PyMuPDF: {e}")
            return "", 0

    def extract_exposure_from_pdf(self, pdf_file_obj: io.BytesIO) -> str:
        try:
            pdf_bytes = pdf_file_obj.getvalue()
            with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
                if doc.page_count == 0:
                    return ""
                page = doc[0]
                search_rect = fitz.Rect(
                    page.rect.width * 0.75, 0, page.rect.width, page.rect.height * 0.15
                )
                text = page.get_text("text", clip=search_rect, sort=True)
                matches = re.findall(r"\b[NSWE]+\b", text.upper())
                if matches:
                    letters = sorted(set("".join(matches)))
                    return ", ".join(letters)
        except Exception:
            pass
        return ""

    def extract_height_in_cm(self, text: str) -> str:
        match = re.search(r"H\s*=\s*(\d+)", text, re.IGNORECASE)
        return match.group(1).strip() if match else ""

    def extract_info_from_filename(self, filename: str) -> Tuple[Optional[str], Optional[str]]:
        try:
            base_name = os.path.splitext(filename)[0]
            normalized = base_name.replace("\u2013", "-").replace("\u2014", "-").strip()
            apt_matches = re.findall(r"\b(\d+\s*[A-Za-z]{1,2})\b", normalized)
            apartment_code: Optional[str] = None
            if apt_matches:
                apartment_code = re.sub(r"\s+", "", apt_matches[-1])
            else:
                tail = normalized.split("_")[-1].split("-")[-1]
                m = re.search(r"(\d+\s*[A-Za-z]{1,2})", tail)
                if m:
                    apartment_code = re.sub(r"\s+", "", m.group(1))

            building_num: Optional[str] = None
            if apartment_code:
                m2 = re.match(r"^(\d+)", apartment_code)
                if m2:
                    building_num = m2.group(1)

            if building_num or apartment_code:
                return building_num, apartment_code
        except Exception:
            return None, None
        return None, None

    def extract_apartment_code_from_text(self, text: str) -> Optional[str]:
        matches = re.findall(r"Oznaczenie lokalu[:\s]+([\w\d/.-]+)", text, re.IGNORECASE)
        for match in matches:
            if not re.match(r"\d{2}-\d{3}", match):
                return match.strip()
        return None

    def extract_building_number_from_text(self, text: str, apartment_code: Optional[str]) -> Optional[str]:
        if apartment_code:
            match = re.search(r"^(\d+)", apartment_code)
            if match:
                return match.group(1)
        return None

    def _extract_value(self, text: str, patterns: List[str], default_value: str = "") -> str:
        normalized_text = re.sub(r"\s+", " ", text).replace(",", ".")
        for pattern in patterns:
            full_pattern = pattern + r"[:\s.a-zA-Z]*?(\d+\.?\d*)"
            matches = re.findall(full_pattern, normalized_text, re.IGNORECASE)
            if matches:
                return matches[-1].strip()
        return default_value

    def extract_surface_area(self, text: str, keywords: List[str]) -> str:
        return self._extract_value(text, keywords)

    def extract_floor_areas(self, text: str) -> Dict[str, str]:
        areas = {
            "parter_uzytkowa": "",
            "pietro_uzytkowa": "",
            "pietro2_uzytkowa": "",
            "parter_scianki": "",
            "pietro_scianki": "",
            "pietro2_scianki": "",
            "laczna": "",
        }
        patterns_uzytkowa = [r"POWIERZCHNIA UŻYTKOWA"]
        patterns_scianki = [r"POWIERZCHNIA POD ŚCIANKAMI(?: I SCHODÓW)?"]
        patterns_laczna = [r"Razem Łączna Powierzchnia Lokalu", r"Łączna Powierzchnia Lokalu"]
        sections = re.split(r"--- STRONA \d+ ---", text)

        if len(sections) > 1:
            areas["parter_uzytkowa"] = self._extract_value(sections[1], patterns_uzytkowa)
            areas["parter_scianki"] = self._extract_value(sections[1], patterns_scianki)
        if len(sections) > 2:
            areas["pietro_uzytkowa"] = self._extract_value(sections[2], patterns_uzytkowa)
            areas["pietro_scianki"] = self._extract_value(sections[2], patterns_scianki)
        if len(sections) > 3:
            areas["pietro2_uzytkowa"] = self._extract_value(sections[3], patterns_uzytkowa)
            areas["pietro2_scianki"] = self._extract_value(sections[3], patterns_scianki)

        areas["laczna"] = self._extract_value(text, patterns_laczna)
        return areas

    def process_pdf(self, pdf_file_obj: io.BytesIO, filename: str) -> Tuple[Dict[str, str], str]:
        exposure = self.extract_exposure_from_pdf(pdf_file_obj)
        pdf_file_obj.seek(0)

        text, page_count = self.extract_text_from_pdf(pdf_file_obj)
        if not text:
            return self.empty_row(), ""

        apartment_code = self.extract_apartment_code_from_text(text)
        if not apartment_code:
            _, apartment_code = self.extract_info_from_filename(filename)

        floor_areas = self.extract_floor_areas(text)
        parter_uzytkowa = floor_areas.get("parter_uzytkowa", "")
        pietro_uzytkowa = floor_areas.get("pietro_uzytkowa", "")
        pietro2_uzytkowa = floor_areas.get("pietro2_uzytkowa", "")
        parter_scianki = floor_areas.get("parter_scianki", "")
        pietro_scianki = floor_areas.get("pietro_scianki", "")
        pietro2_scianki = floor_areas.get("pietro2_scianki", "")

        laczna_parter = self._safe_float_sum(parter_uzytkowa, parter_scianki)
        laczna_pietro = self._safe_float_sum(pietro_uzytkowa, pietro_scianki)
        laczna_pietro2 = self._safe_float_sum(pietro2_uzytkowa, pietro2_scianki)
        laczna_uzytkowa = self._safe_float_sum(parter_uzytkowa, pietro_uzytkowa, pietro2_uzytkowa)

        garden_area = self.extract_surface_area(text, [r"Powierzchnia ogrodu", r"ogród"])

        sections = re.split(r"--- STRONA \d+ ---", text)
        parter_height_cm = self.extract_height_in_cm(sections[1]) if len(sections) > 1 else ""
        pietro_height_cm = self.extract_height_in_cm(sections[2]) if len(sections) > 2 else ""
        poddasze_height_cm = self.extract_height_in_cm(sections[3]) if len(sections) > 3 else ""

        parter_height_m = f"{float(parter_height_cm) / 100:.2f}" if parter_height_cm.isdigit() else ""
        pietro_height_m = f"{float(pietro_height_cm) / 100:.2f}" if pietro_height_cm.isdigit() else ""
        poddasze_height_m = (
            f"{float(poddasze_height_cm) / 100:.2f}" if poddasze_height_cm.isdigit() else ""
        )

        def present(patterns: List[str]) -> str:
            return "1" if any(re.search(p, text, re.IGNORECASE) for p in patterns) else "0"

        if bool(re.search(r"\bsalon\s+z\s+(kuchni|aneks)\w*", text, re.IGNORECASE)):
            has_kuchnia, has_aneks = "0", "1"
        else:
            has_aneks = present([r"\baneks\w*\s+kuchenn\w*"])
            has_kuchnia = "0" if has_aneks == "1" else present([r"\bkuchni\w*"])

        rooms_matches = re.findall(
            r"L\.p\.[\s\S]{0,300}?Pomieszczenie([\s\S]*?)RAZEM", text, re.IGNORECASE
        )
        bathrooms_count = 0
        rooms_count = 0
        if rooms_matches:
            for rooms_block in rooms_matches:
                bathrooms_count += len(
                    re.findall(r"\d+\.\d+\s+(?:ł|l)azienk\w*", rooms_block, re.IGNORECASE)
                )
                rooms_count += len(
                    re.findall(
                        r"\d+\.\d+\s+(?:salon|sypialn\w*|gabine\w*|pok[oó]j\w*)",
                        rooms_block,
                        re.IGNORECASE,
                    )
                )
        else:
            bathrooms_count = len(re.findall(r"\d+\.\d+\s+(?:ł|l)azienk\w*", text, re.IGNORECASE))
            rooms_count = len(
                re.findall(
                    r"\d+\.\d+\s+(?:salon|sypialn\w*|gabine\w*|pok[oó]j\w*)", text, re.IGNORECASE
                )
            )

        result = self.empty_row()
        ocr_values = {
            "NUMER LOKALU": apartment_code or "",
            "EKSPOZYCJA ": exposure,
            "POWIERZCHNIA UZYTKOWA PARTER": parter_uzytkowa,
            "POWIERZCHNIA UZYTKOWA PIĘTRO ": pietro_uzytkowa,
            "POWIERZCHNIA UZYTKOWA PIETRO 2": pietro2_uzytkowa if page_count >= 3 else "",
            "Powierzchnia pod ściankami i schodów Parter ": parter_scianki,
            "Powierzchnia pod ściankami  i schodów  Piętro  ": pietro_scianki,
            "Powierzchnia pod ściankami  i schodów  Piętro  2": pietro2_scianki if page_count >= 3 else "",
            "łączna powierzchnia parter": laczna_parter,
            "łączna powierzchnia piętro": laczna_pietro,
            "łączna powierzchnia piętro 2": laczna_pietro2 if page_count >= 3 else "",
            "POWIERZCHNIA ŁĄCZNIE": floor_areas.get("laczna", ""),
            "POWIERZCHNIA UŻYTKOWA ŁĄCZNIE": laczna_uzytkowa,
            "LICZBA POKOI ": str(rooms_count) if rooms_count > 0 else "",
            "LICZBA KONDYGNACJI ": str(page_count) if page_count > 0 else "",
            "OGRÓD": garden_area,
            "PODDASZE": "1" if page_count >= 3 else "0",
            "Wysokość PARTERU": parter_height_m,
            "Wysokość PIĘTRA": pietro_height_m,
            "Wysokość PODDASZA": poddasze_height_m if page_count >= 3 else "",
            "TARAS ": present([r"\btaras\w*"]),
            "LOGGIA": present([r"\bloggi\w*"]),
            "BALKON": present([r"\bbalkon\w*"]),
            "KUCHNIA": has_kuchnia,
            "ANEKS KUCHENNY ": has_aneks,
            "GARDEROBA": present([r"\bgarderob\w*"]),
            "KOMINEK ": present([r"\bkomine\w*"]),
            "Schowek": present([r"\bschow\w*"]),
            "Spiżarnia": present([r"\bspi(?:ż|z)arni\w*"]),
            "Pomieszczenie gospodarcze": present(
                [r"\bpom\.?\s*gosp\.?\b", r"\bpomieszczenie\s+gospodarcze\b"]
            ),
            "Druga łazienka z prysznicem": "1" if bathrooms_count >= 2 else "0",
            "Gabinet": present([r"\bgabine\w*"]),
        }
        result.update(ocr_values)
        return result, text


# --- Interfejs Użytkownika Streamlit ---

st.set_page_config(layout="wide")
st.title("OCR do ekstrakcji danych — format Excel ostateczny")
st.subheader("PyMuPDF + eksport zgodny z szablonem (tylko kolumny OCR)")
st.write(
    "Wgraj pliki PDF. Program wypełni wyłącznie kolumny oznaczone jako **ocr**. "
    "Kolumny **marketing** i **architekt** pozostają puste — uzupełniacie je ręcznie."
)

ocr = ApartmentPlanOCR()

uploaded_files = st.file_uploader(
    "Wybierz pliki PDF (np. 'OP6_ZT-1A.pdf')",
    type="pdf",
    accept_multiple_files=True,
)

if uploaded_files:
    if st.button("Przetwórz pliki"):
        all_results = []
        progress_bar = st.progress(0, text="Oczekiwanie...")

        with st.spinner("Analizuję dokumenty..."):
            for i, uploaded_file in enumerate(uploaded_files):
                progress_bar.progress(
                    (i + 1) / len(uploaded_files), text=f"Przetwarzanie: {uploaded_file.name}"
                )
                result, full_text = ocr.process_pdf(uploaded_file, uploaded_file.name)
                all_results.append(result)

                if full_text:
                    with st.expander(f"Pokaż/Ukryj pełny tekst odczytany z '{uploaded_file.name}'"):
                        st.text_area("Surowy tekst OCR:", full_text, height=300, key=f"text_area_{i}")
                else:
                    st.warning(f"Nie udało się odczytać tekstu z pliku '{uploaded_file.name}'.")

        if all_results:
            st.markdown("---")
            st.header("Zbiorcze wyniki (wszystkie kolumny)")
            st.success(f"Zakończono! Przetworzono {len(all_results)} plików.")
            st.caption("Kolumny marketing i architekt są puste — uzupełnicie je ręcznie.")
            df_results = results_to_dataframe(all_results)
            st.dataframe(df_results)
            st.session_state["df_results"] = all_results
        else:
            st.warning("Nie udało się wydobyć danych z żadnego z plików.")

if "df_results" in st.session_state:
    excel_data = export_results_to_excel(st.session_state["df_results"])

    st.download_button(
        label="Pobierz wyniki jako plik Excel (format szablonu)",
        data=excel_data,
        file_name="wyniki_ocr.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheet.sheet",
    )
